import uuid
from io import BytesIO
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from datetime import datetime

from domain.contracts.core.file_download_url_contract import FileDownloadUrlContract
from domain.contracts.core.international_offer_contract import (
    InternationalOfferDetailContract,
    InternationalOfferItemContract,
    InternationalOfferContract,
)
from domain.contracts.core.national_offer_contract import (
    NationalOfferContract,
    NationalOfferDetailContract,
    NationalOfferLocationContract,
)
from domain.models.core.offer_location_model import VwOfferLocationModel
from domain.models.core.offer_model import (
    VwOfferNationalModel,
    VwOfferInternationalModel,
)
from domain.models.operation.operation_detail_model import VwOperationDetailModel
from domain.models.operation.operation_model import VwOperationModel
from domain.models.supplier.supplier_operation_model import SupplierOperationModel
from domain.services.core.offer_service import OfferService
from domain.services.operation.operation_service import OperationService
from domain.services.supplier.supplier_operation_service import SupplierOperationService
from repository.core.offer_repository import OfferRepository
from openpyxl import load_workbook


class TestOfferService:
    @pytest.fixture
    def service(self) -> OfferService:
        return OfferService()

    @pytest.fixture
    def repository(self) -> OfferRepository:
        return OfferRepository()

    async def test_get_offer_national_by_operation_id_returns_expected_data_with_locations(
        self, service: OfferService, repository: OfferRepository
    ):
        # Arrange
        mock_data = [
            {
                "offer_id": 1,
                "supplier_name": "Proveedor S.A.",
                "supplier_id": 100,
                "currency_name": "USD",
                "free_loading": True,
                "offer_code": "OP-001-1",
                "valid_until": datetime(2025, 5, 20),
                "load_date": datetime(2025, 5, 15),
                "total_location_tariff": 1500.0,
                "locations_detail": [
                    {
                        "location": "Lima",
                        "address": "Av. Principal 123",
                        "tariff": 750.0,
                    },
                    {"location": "Callao", "address": "Puerto 456", "tariff": 750.0},
                ],
                "selected": False,
            }
        ]
        service.repository = repository
        service.repository.get_vw_offer_national_by_operation_id = AsyncMock(
            return_value=mock_data
        )

        # Act
        result = await service.get_offer_national_by_operation_id(operation_id=123)

        # Assert
        assert isinstance(result, list)
        assert len(result) == 1

        offer = result[0]
        assert isinstance(offer, VwOfferNationalModel)
        assert isinstance(offer.locations_detail, list)
        assert len(offer.locations_detail) == 2

        loc1 = offer.locations_detail[0]
        assert isinstance(loc1, VwOfferLocationModel)

        service.repository.get_vw_offer_national_by_operation_id.assert_called_once_with(
            123
        )

    @pytest.mark.asyncio
    async def test_get_offer_international_by_operation_id_returns_expected_data(
        self, service: OfferService, repository: OfferRepository
    ):
        # Arrange
        mock_data = [
            {
                "operation_id": 123,
                "offer_id": 1,
                "location": "Ubicacion",
                "address": "Puerto Internacional",
                "supplier_name": "Global Freight GmbH",
                "supplier_id": 42,
                "offer_code": "INT-123-1",
                "valid_until": datetime(2025, 6, 1),
                "load_date": "2025-05-15",
                "tariff": 3200.5,
                "currency_name": "EUR",
                "incoterm_code": "FOB",
                "transit_time": 15,
                "transshipment": False,
                "free_days": 7,
                "transporter": "Maestre",
                "min": 1000.0,
                "amount": 2500.0,
                "wm": 12.5,
                "imo": None,
                "container_size_20": 1200.0,
                "container_size_40": 2000.0,
                "freight_percentage": 5.5,
                "document_bl": 100.0,
                "preparation_fee_bl": 50.0,
                "mounting_dismounting": 0.0,
                "food_grade": 20.0,
                "positioning": 30.0,
                "thc_origin": 150.0,
                "special_handling": 80.0,
                "vgm": 10.0,
                "custom_ams": 60.0,
                "consolidation_lcl": 90.0,
                "destination_bl": 40.0,
                "destination_cont": 70.0,
                "free_days_destination": 5,
                "selected": False,
            }
        ]

        service.repository = repository
        service.repository.get_vw_offer_international_by_operation_id = AsyncMock(
            return_value=mock_data
        )

        # Act
        result = await service.get_offer_international_by_operation_id(123)

        # Assert
        assert isinstance(result, list)
        assert len(result) == 1

        offer = result[0]
        assert isinstance(offer, VwOfferInternationalModel)

        service.repository.get_vw_offer_international_by_operation_id.assert_called_once_with(
            123
        )

    @pytest.mark.asyncio
    async def test_get_offer_national_by_operation_offer_id_returns_expected_model(
        self, service: OfferService, repository: OfferRepository
    ):
        # Arrange
        mock_data = {
            "offer_id": 1,
            "supplier_name": "Proveedor Nacional",
            "supplier_id": 100,
            "currency_name": "PEN",
            "free_loading": False,
            "offer_code": "OP-456-1",
            "valid_until": datetime(2025, 7, 1),
            "load_date": datetime(2025, 6, 20),
            "total_location_tariff": 1000.0,
            "locations_detail": [
                {"location": "Arequipa", "address": "Zona Industrial", "tariff": 500.0},
                {"location": "Cusco", "address": "Av. Los Incas", "tariff": 500.0},
            ],
            "selected": True,
        }

        service.repository = repository
        service.repository.get_vw_offer_national_by_operation_offer_id = AsyncMock(
            return_value=mock_data
        )

        # Act
        result = await service.get_offer_national_by_operation_offer_id(
            operation_id=200, offer_id=1
        )

        # Assert
        assert isinstance(result, VwOfferNationalModel)
        service.repository.get_vw_offer_national_by_operation_offer_id.assert_called_once_with(
            200, 1
        )

    @pytest.mark.asyncio
    async def test_get_offer_international_by_operation_offer_id_returns_expected_data(
        self, service: OfferService, repository: OfferRepository
    ):
        # Arrange
        mock_data = [
            {
                "operation_id": 300,
                "offer_id": 2,
                "location": "Rotterdam",
                "address": "Dock A1",
                "supplier_name": "Global Logistics",
                "supplier_id": 55,
                "offer_code": "INT-456-2",
                "valid_until": datetime(2025, 8, 10),
                "load_date": "2025-08-01",
                "tariff": 4200.0,
                "currency_name": "USD",
                "incoterm_code": "CIF",
                "transit_time": 20,
                "transshipment": True,
                "free_days": 10,
                "transporter": "MSC",
                "min": 1500.0,
                "amount": 3200.0,
                "wm": 14.0,
                "imo": 10.0,
                "container_size_20": 1300.0,
                "container_size_40": 2100.0,
                "freight_percentage": 6.0,
                "document_bl": 120.0,
                "preparation_fee_bl": 60.0,
                "mounting_dismounting": 10.0,
                "food_grade": 15.0,
                "positioning": 25.0,
                "thc_origin": 160.0,
                "special_handling": 70.0,
                "vgm": 15.0,
                "custom_ams": 55.0,
                "consolidation_lcl": 85.0,
                "destination_bl": 45.0,
                "destination_cont": 75.0,
                "free_days_destination": 6,
                "selected": True,
            }
        ]

        service.repository = repository
        service.repository.get_vw_offer_international_by_operation_offer_id = AsyncMock(
            return_value=mock_data
        )

        # Act
        result = await service.get_offer_international_by_operation_offer_id(
            operation_id=300, offer_id=2
        )

        # Assert
        assert isinstance(result, list)
        assert len(result) == 1
        assert isinstance(result[0], VwOfferInternationalModel)

        service.repository.get_vw_offer_international_by_operation_offer_id.assert_called_once_with(
            300, 2
        )

    @pytest.mark.asyncio
    async def test_create_offer_national_creates_offers_correctly(
        self, service: OfferService, repository: OfferRepository
    ):
        # Arrange
        service.repository = repository
        service.supplier_operation_service = SupplierOperationService()
        service.operation_service = OperationService()
        service.repository = MagicMock()
        service.offer_location_service = MagicMock()
        service.supplier_operation_email_service = MagicMock()
        service.operation_notification_service = MagicMock()

        key = uuid.uuid4()

        supplier_operation_mock = SupplierOperationModel(
            supplier_operation_id=1,
            operation_id=1,
            supplier_id=45,
            key=key,
        )

        operation_mock = MagicMock(operation_code="OP123", offer_count=2)

        service.supplier_operation_service.get_operation_by_key = AsyncMock(
            return_value=supplier_operation_mock
        )
        service.operation_service.get_vw_operation_by_operation_id = AsyncMock(
            return_value=operation_mock
        )
        service.repository.create = AsyncMock(return_value=100)
        service.offer_location_service.bulk_create = AsyncMock()
        service.supplier_operation_email_service.get_notified_email_by_operation_supplier = AsyncMock(
            return_value=["test@example.com"]
        )
        service.operation_notification_service.send_new_offer_national = AsyncMock()

        offer_location = NationalOfferLocationContract(
            location_id=1, location_name="Lima", address="Av. Principal", tariff=500.0
        )
        offer_detail = NationalOfferDetailContract(
            valid_until=datetime(2025, 5, 30),
            free_loading=True,
            availability_requested_date=False,
            load_date=datetime(2025, 5, 20),
            locations=[offer_location],
            total_tariff=500.0,
            currency_id=1,
            currency_name="USD",
            observation="Test observation",
        )
        contract = NationalOfferContract(quotations=[offer_detail])

        # Act
        await service.create_offer_national(contract, key)

        # Assert
        service.supplier_operation_service.get_operation_by_key.assert_called_once_with(
            key
        )
        service.operation_service.get_vw_operation_by_operation_id.assert_called_once_with(
            1
        )
        service.repository.create.assert_called_once()
        service.offer_location_service.bulk_create.assert_called_once()
        service.supplier_operation_email_service.get_notified_email_by_operation_supplier.assert_called_once_with(
            1, 45
        )
        service.operation_notification_service.send_new_offer_national.assert_called_once()

    async def test_create_offer_international_creates_all_offers_and_sends_notification(
        self, service: OfferService, repository: OfferRepository
    ):
        # Arrange
        key = uuid.uuid4()

        supplier_operation_mock = SupplierOperationModel(
            supplier_operation_id=1,
            operation_id=99,
            supplier_id=50,
            key=key,
        )
        operation_detail_mock = [
            VwOperationDetailModel(
                unit_count=10,
                unit_weight=1500.0,
                net_weight=1400.0,
                gross_weight=1550.0,
                height=2.5,
                width=2.0,
                length=6.0,
                unit_type="Container",
                location="Puerto de Medellin",
            ),
            VwOperationDetailModel(
                unit_count=5,
                unit_weight=2500.0,
                net_weight=2400.0,
                gross_weight=2600.0,
                height=2.8,
                width=2.5,
                length=6.5,
                unit_type="Camión",
                location="Puerto de Rotterdam",
            ),
        ]

        operation_mock = VwOperationModel(
            operation_id=123,
            operation_code="OP-12345",
            user_id=42,
            requires_escort=True,
            load_date=datetime(2025, 5, 20),
            created_at=datetime(2025, 1, 1),
            observation="Operación internacional",
            validity_date=datetime(2025, 6, 30),
            operation_type_name="Internacional",
            email="cliente@empresa.com",
            location_origin_name="Puerto de Medellin",
            service_type_name="Marítimo",
            operation_status_name="Pendiente",
            incoterms=["FOB", "CIF"],
            offer_count=5,
            operation_details=operation_detail_mock,
        )

        service.supplier_operation_service.get_operation_by_key = AsyncMock(
            return_value=supplier_operation_mock
        )
        service.operation_service.get_vw_operation_by_operation_id = AsyncMock(
            return_value=operation_mock
        )
        service.repository.create = AsyncMock(side_effect=[1001, 1002])
        service.offer_location_service.bulk_create = AsyncMock()
        service.supplier_operation_email_service.get_notified_email_by_operation_supplier = AsyncMock(
            return_value=["notify@export.com"]
        )
        service.operation_notification_service.send_new_offer_international = (
            AsyncMock()
        )

        quotation1 = InternationalOfferDetailContract(
            incoterm_id=1,
            currency_id=1,
            currency_name="USD",
            valid_until=datetime(2025, 6, 30),
            transit_time=20,
            transshipment=False,
            free_days=5,
            tariff=3500.0,
            free_loading=True,
            observation="Urgente",
            transporter="MSC",
            min=100.0,
            amount=2500.0,
            wm=12.5,
            imo=0.0,
            container_size_20=1000.0,
            container_size_40=1500.0,
            freight_percentage=5.0,
            document_bl=50.0,
            preparation_fee_bl=20.0,
            mounting_dismounting=0.0,
            food_grade=10.0,
            positioning=30.0,
            thc_origin=150.0,
            special_handling=40.0,
            vgm=10.0,
            custom_ams=25.0,
            consolidation_lcl=30.0,
            destination_bl=15.0,
            destination_cont=40.0,
            free_days_destination=3,
        )

        quotation2 = InternationalOfferDetailContract(
            incoterm_id=2,
            currency_id=1,
            currency_name="USD",
            valid_until=datetime(2025, 7, 15),
            transit_time=25,
            transshipment=True,
            free_days=7,
            tariff=3800.0,
            free_loading=False,
            observation="Normal",
            transporter="Medellin",
            min=200.0,
            amount=3000.0,
            wm=14.0,
            imo=1.0,
            container_size_20=1100.0,
            container_size_40=1600.0,
            freight_percentage=6.0,
            document_bl=60.0,
            preparation_fee_bl=25.0,
            mounting_dismounting=5.0,
            food_grade=15.0,
            positioning=35.0,
            thc_origin=160.0,
            special_handling=50.0,
            vgm=12.0,
            custom_ams=30.0,
            consolidation_lcl=35.0,
            destination_bl=20.0,
            destination_cont=50.0,
            free_days_destination=4,
        )

        location_contract = InternationalOfferItemContract(
            location_id=10,
            location_name="Rotterdam",
            address="Puerto 101",
            quotations=[quotation1, quotation2],
        )

        contract = InternationalOfferContract(locations=[location_contract])

        # Act
        await service.create_offer_international(contract, key)

        # Assert
        service.supplier_operation_service.get_operation_by_key.assert_called_once_with(
            key
        )
        service.operation_service.get_vw_operation_by_operation_id.assert_called_once_with(
            99
        )
        assert service.repository.create.call_count == 2
        service.offer_location_service.bulk_create.assert_called_once()
        service.supplier_operation_email_service.get_notified_email_by_operation_supplier.assert_called_once_with(
            99, 50
        )
        service.operation_notification_service.send_new_offer_international.assert_called_once_with(
            contract, ["notify@export.com"], "OP-12345"
        )

    @pytest.mark.asyncio
    async def test_notified_offer_selected_international(
        self, service: OfferService, repository: OfferRepository
    ):
        # Arrange
        operation_id = 1
        offer_id = 123
        mock_operation = VwOperationModel(
            operation_id=operation_id,
            operation_code="OP-12345",
            user_id=42,
            requires_escort=False,
            load_date=datetime(2025, 5, 20),
            created_at=datetime(2025, 1, 1),
            observation="Operación internacional",
            validity_date=datetime(2025, 6, 30),
            operation_type_name="Internacional",
            email="operacion@empresa.com",
            location_origin_name="Puerto Internacional",
            service_type_name="Marítimo",
            operation_status_name="Pendiente",
            incoterms=["FOB"],
            offer_count=5,
            operation_details=[],
        )

        mock_offer = VwOfferInternationalModel(
            operation_id=operation_id,
            offer_id=offer_id,
            location="Lugar A",
            address="Dirección A",
            supplier_name="Proveedor A",
            supplier_id=42,
            offer_code="OP-12345-6",
            valid_until=datetime(2025, 6, 1),
            load_date="2025-05-20",
            tariff=1000.0,
            currency_name="USD",
            incoterm_code="FOB",
            transit_time=10,
            transshipment=False,
            free_days=5,
            transporter="Transportista A",
            min=100.0,
            amount=5000.0,
            wm=100.0,
            imo=12345.0,
            container_size_20=20.0,
            container_size_40=40.0,
            freight_percentage=10.0,
            document_bl=50.0,
            preparation_fee_bl=20.0,
            mounting_dismounting=15.0,
            food_grade=10.0,
            positioning=5.0,
            thc_origin=3.0,
            special_handling=8.0,
            vgm=500.0,
            custom_ams=25.0,
            consolidation_lcl=12.0,
            destination_bl=30.0,
            destination_cont=40.0,
            free_days_destination=3,
            selected=False,
        )

        service.operation_service.get_vw_operation_by_operation_id = AsyncMock(
            return_value=mock_operation
        )
        service.get_offer_international_by_operation_offer_id = AsyncMock(
            return_value=[mock_offer]
        )
        service.supplier_operation_email_service.get_notified_email_by_operation_supplier = AsyncMock(
            return_value=["supplier@example.com"]
        )
        service.offer_notification_service.send_new_confirmation_offer_international = (
            AsyncMock()
        )
        service.supplier_operation_email_service.get_notified_email_by_operation = (
            AsyncMock(return_value=["admin@example.com"])
        )
        service.offer_notification_service.send_closed_operation = AsyncMock()
        service.mark_offer_as_selected = AsyncMock()
        service.operation_service.update_status_operation = AsyncMock()
        # Act
        result = await service.notified_offer_selected_international(
            operation_id, offer_id
        )

        # Assert
        service.operation_service.get_vw_operation_by_operation_id.assert_called_once_with(
            operation_id
        )
        service.get_offer_international_by_operation_offer_id.assert_called_once_with(
            operation_id, offer_id
        )
        service.supplier_operation_email_service.get_notified_email_by_operation_supplier.assert_called_once_with(
            operation_id, mock_offer.supplier_id
        )
        service.offer_notification_service.send_new_confirmation_offer_international.assert_called_once_with(
            [mock_offer], ["supplier@example.com"], mock_operation
        )
        service.offer_notification_service.send_closed_operation.assert_called_once_with(
            ["admin@example.com"], mock_operation
        )
        service.supplier_operation_email_service.get_notified_email_by_operation.assert_called_once_with(
            operation_id
        )
        service.mark_offer_as_selected.assert_called_once_with(offer_id)
        assert result is True

    @pytest.mark.asyncio
    async def test_notified_offer_selected_national(
        self, service: OfferService, repository: OfferRepository
    ):
        # Arrange
        operation_id = 1
        offer_id = 123
        mock_operation = VwOperationModel(
            operation_id=operation_id,
            operation_code="OP-12345",
            user_id=42,
            requires_escort=False,
            load_date=datetime(2025, 5, 20),
            created_at=datetime(2025, 1, 1),
            observation="Operación nacional",
            validity_date=datetime(2025, 6, 30),
            operation_type_name="Nacional",
            email="operacion@empresa.com",
            location_origin_name="Puerto Nacional",
            service_type_name="Terrestre",
            operation_status_name="Pendiente",
            incoterms=["FOB"],
            offer_count=5,
            operation_details=[],
        )

        mock_offer = VwOfferNationalModel(
            offer_id=offer_id,
            supplier_name="Proveedor A",
            supplier_id=42,
            currency_name="MXN",
            free_loading=True,
            offer_code="OP-12345-6",
            valid_until=datetime(2025, 6, 1),
            load_date=datetime(2025, 5, 20),
            total_location_tariff=1000.0,
            locations_detail=[
                VwOfferLocationModel(
                    location="Lugar A", address="Dirección A", tariff=500.0
                )
            ],
            selected=False,
        )

        service.operation_service.get_vw_operation_by_operation_id = AsyncMock(
            return_value=mock_operation
        )
        service.get_offer_national_by_operation_offer_id = AsyncMock(
            return_value=mock_offer
        )
        service.supplier_operation_email_service.get_notified_email_by_operation_supplier = AsyncMock(
            return_value=["supplier@example.com"]
        )
        service.offer_notification_service.send_new_confirmation_offer_national = (
            AsyncMock()
        )
        service.supplier_operation_email_service.get_notified_email_by_operation = (
            AsyncMock(return_value=["admin@example.com"])
        )
        service.offer_notification_service.send_closed_operation = AsyncMock()
        service.mark_offer_as_selected = AsyncMock()
        service.operation_service.update_status_operation = AsyncMock()
        # Act
        result = await service.notified_offer_selected_national(operation_id, offer_id)

        # Assert
        service.operation_service.get_vw_operation_by_operation_id.assert_called_once_with(
            operation_id
        )
        service.get_offer_national_by_operation_offer_id.assert_called_once_with(
            operation_id, offer_id
        )
        service.supplier_operation_email_service.get_notified_email_by_operation_supplier.assert_called_once_with(
            operation_id, mock_offer.supplier_id
        )
        service.offer_notification_service.send_new_confirmation_offer_national.assert_called_once_with(
            mock_offer, ["supplier@example.com"], mock_operation
        )
        service.supplier_operation_email_service.get_notified_email_by_operation.assert_called_once_with(
            operation_id
        )
        service.offer_notification_service.send_closed_operation.assert_called_once_with(
            ["admin@example.com"], mock_operation
        )

        service.mark_offer_as_selected.assert_called_once_with(offer_id)
        assert result is True

    @patch("domain.services.core.offer_service.generate_presigned_url")
    @patch("domain.services.core.offer_service.upload_file")
    async def test_download_national_offers_excel(
        self, upload_file_mock: MagicMock, generate_url_mock: MagicMock, service
    ):
        # Arrange
        offer = [
            VwOfferNationalModel(
                offer_id=1,
                supplier_name="Supplier A",
                supplier_id=123,
                currency_name="USD",
                free_loading=True,
                offer_code="OFF123",
                valid_until=datetime(2025, 12, 31),
                load_date=None,
                total_location_tariff=1500.0,
                selected=True,
                locations_detail=[
                    VwOfferLocationModel(
                        location="Lima", address="Av. Ejemplo 123", tariff=100.0
                    ),
                    VwOfferLocationModel(
                        location="Arequipa", address="Av. Ejemplo 456", tariff=120.0
                    ),
                ],
            )
        ]

        service.get_offer_national_by_operation_id = AsyncMock(return_value=offer)

        url = "https://example.com/download/offers.xlsx"
        generate_url_mock.return_value = url
        captured = {}

        def fake_upload(file_path, s3_path, remove_on_upload=True):
            with open(file_path, "rb") as f:
                captured["data"] = f.read()

        upload_file_mock.side_effect = fake_upload
        # Act
        result = await service.download_national_offers_excel(1)

        # Assert
        wb = load_workbook(BytesIO(captured["data"]))
        ws = wb.active
        assert ws["A2"].value == "Supplier A"
        service.get_offer_national_by_operation_id.assert_awaited_once_with(1)
        upload_file_mock.assert_called_once()
        generate_url_mock.assert_called_once()

        assert isinstance(result, FileDownloadUrlContract)
        assert result.url == url

    @patch("domain.services.core.offer_service.generate_presigned_url")
    @patch("domain.services.core.offer_service.upload_file")
    async def test_download_international_offers_excel(
        self, upload_file_mock: MagicMock, generate_url_mock: MagicMock, service
    ):
        # Arrange
        mock_offer = VwOfferInternationalModel(
            operation_id=1,
            offer_id=100,
            location="Medellin",
            address="Port German 123",
            supplier_name="Proveedor B",
            supplier_id=42,
            offer_code="INTL-OFFER-001",
            valid_until=datetime(2025, 12, 31),
            load_date="2025-11-01",
            tariff=2500.0,
            currency_name="EUR",
            incoterm_code="FOB",
            transit_time=20,
            transshipment=False,
            free_days=7,
            transporter="Samuel-Lloyd",
            min=100.0,
            amount=1200.0,
            wm=1.5,
            imo=300.0,
            container_size_20=1200.0,
            container_size_40=2300.0,
            freight_percentage=10.0,
            document_bl=50.0,
            preparation_fee_bl=75.0,
            mounting_dismounting=100.0,
            food_grade=20.0,
            positioning=30.0,
            thc_origin=200.0,
            special_handling=90.0,
            vgm=15.0,
            custom_ams=25.0,
            consolidation_lcl=60.0,
            destination_bl=80.0,
            destination_cont=160.0,
            free_days_destination=5,
            selected=True,
        )

        service.get_offer_international_by_operation_id = AsyncMock(
            return_value=[mock_offer]
        )
        url = "https://example.com/download/offers.xlsx"
        generate_url_mock.return_value = url

        captured = {}

        def fake_upload(file_path, s3_path, remove_on_upload=True):
            with open(file_path, "rb") as f:
                captured["data"] = f.read()

        upload_file_mock.side_effect = fake_upload

        # Act
        result = await service.download_international_offers_excel(1)

        # Assert
        wb = load_workbook(BytesIO(captured["data"]))
        ws = wb.active

        assert ws.title == "Offers"
        service.get_offer_international_by_operation_id.assert_awaited_once_with(1)
        upload_file_mock.assert_called_once()
        generate_url_mock.assert_called_once()
        assert isinstance(result, FileDownloadUrlContract)
        assert result.url == url
